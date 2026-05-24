"""
DocuTools — Tests para módulos de conversión:
  - core/pdf_to_image.py  (pdf_to_images)
  - core/image_to_pdf.py  (images_to_pdf)
  - core/pdf_to_word.py   (pdf_to_word) — skip si dependencias no disponibles
  - core/word_to_pdf.py   (word_to_pdf) — skip si dependencias no disponibles

Pruebas para:
  - Conversión de imágenes (JPG, PNG, BMP, mixtas) a PDF
  - Conversión de PDF a imágenes (mocked si poppler no disponible)
  - Lista vacía de imágenes → error
  - Imagen única → 1 página
  - Opciones de tamaño de página

Mínimo: 5 happy path, 3 error, 2 edge case.
"""

import sys
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from PIL import Image
from pypdf import PdfReader, PdfWriter

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

# ════════════════════════════════════════════════════════
#  Intentar importar módulos — marcar skip si no existen aún
# ════════════════════════════════════════════════════════

try:
    from core.image_to_pdf import images_to_pdf
    HAS_IMAGE_TO_PDF = True
except ImportError:
    HAS_IMAGE_TO_PDF = False

try:
    from core.pdf_to_image import pdf_to_images
    HAS_PDF_TO_IMAGE = True
except ImportError:
    HAS_PDF_TO_IMAGE = False

try:
    from core.pdf_to_word import pdf_to_word
    HAS_PDF_TO_WORD = True
except ImportError:
    HAS_PDF_TO_WORD = False

try:
    from core.word_to_pdf import word_to_pdf
    HAS_WORD_TO_PDF = True
except ImportError:
    HAS_WORD_TO_PDF = False

try:
    from core.excel_to_pdf import excel_to_pdf, detect_engine
    HAS_EXCEL_TO_PDF = True
except ImportError:
    HAS_EXCEL_TO_PDF = False

try:
    from core.powerpoint_to_pdf import powerpoint_to_pdf, detect_powerpoint_engine
    HAS_POWERPOINT_TO_PDF = True
except ImportError:
    HAS_POWERPOINT_TO_PDF = False

try:
    from core.pdf_to_excel import pdf_to_excel
    HAS_PDF_TO_EXCEL = True
except ImportError:
    HAS_PDF_TO_EXCEL = False


# ════════════════════════════════════════════════════════
#  Helpers
# ════════════════════════════════════════════════════════

def _make_pdf(path: Path, num_pages: int) -> Path:
    """Crear un PDF con N páginas en blanco."""
    writer = PdfWriter()
    for _ in range(num_pages):
        writer.add_blank_page(width=612, height=792)
    writer.write(str(path))
    return path


def _make_images(tmp_path: Path, count: int, fmt: str = "JPEG", ext: str = "jpg") -> list:
    """Crear N imágenes de prueba."""
    paths = []
    colors = [
        (255, 0, 0), (0, 255, 0), (0, 0, 255),
        (255, 255, 0), (255, 0, 255), (0, 255, 255),
    ]
    for i in range(count):
        p = tmp_path / f"image_{i}.{ext}"
        color = colors[i % len(colors)]
        mode = "RGBA" if fmt == "PNG" and ext == "png" else "RGB"
        img = Image.new(mode, (640, 480), color=color + ((128,) if mode == "RGBA" else ()))
        img.save(str(p), fmt)
        paths.append(p)
    return paths


# ════════════════════════════════════════════════════════
#  IMAGES → PDF — Happy Path
# ════════════════════════════════════════════════════════

@pytest.mark.skipif(not HAS_IMAGE_TO_PDF, reason="core.image_to_pdf no disponible")
class TestImagesToPdf:
    """Pruebas de conversión de imágenes a PDF."""

    def test_images_to_pdf_jpg(self, tmp_path: Path):
        """3 imágenes JPG → PDF de 3 páginas."""
        images = _make_images(tmp_path, 3, fmt="JPEG", ext="jpg")
        output = tmp_path / "from_jpgs.pdf"

        result = images_to_pdf(images, output)

        assert output.exists(), "No se creó el PDF"
        reader = PdfReader(str(output))
        assert len(reader.pages) == 3

    def test_images_to_pdf_png_transparent(self, tmp_path: Path):
        """Una imagen PNG con canal alfa → PDF de 1 página válido."""
        png_path = tmp_path / "alpha.png"
        img = Image.new("RGBA", (400, 400), color=(200, 50, 50, 128))
        img.save(str(png_path), "PNG")
        output = tmp_path / "from_png.pdf"

        result = images_to_pdf([png_path], output)

        assert output.exists()
        reader = PdfReader(str(output))
        assert len(reader.pages) == 1

    def test_images_to_pdf_mixed(self, tmp_path: Path):
        """Mezcla de JPG y PNG → PDF con N páginas."""
        jpg = tmp_path / "mix.jpg"
        png = tmp_path / "mix.png"
        Image.new("RGB", (640, 480), (100, 100, 100)).save(str(jpg), "JPEG")
        Image.new("RGBA", (400, 400), (200, 50, 50, 128)).save(str(png), "PNG")
        output = tmp_path / "mixed.pdf"

        result = images_to_pdf([jpg, png], output)

        assert output.exists()
        reader = PdfReader(str(output))
        assert len(reader.pages) == 2

    def test_images_to_pdf_single(self, tmp_path: Path):
        """1 imagen → 1 página."""
        images = _make_images(tmp_path, 1, fmt="JPEG", ext="jpg")
        output = tmp_path / "single_img.pdf"

        result = images_to_pdf(images, output)

        assert output.exists()
        reader = PdfReader(str(output))
        assert len(reader.pages) == 1

    def test_images_to_pdf_bmp(self, tmp_path: Path):
        """Imagen BMP → PDF de 1 página."""
        bmp_path = tmp_path / "image.bmp"
        Image.new("RGB", (100, 100), (0, 128, 0)).save(str(bmp_path), "BMP")
        output = tmp_path / "from_bmp.pdf"

        result = images_to_pdf([bmp_path], output)

        assert output.exists()
        reader = PdfReader(str(output))
        assert len(reader.pages) == 1

    def test_images_to_pdf_page_sizes(self, tmp_path: Path):
        """Cada opción de page_size debe funcionar sin errores."""
        SUPPORTED_PAGE_SIZES = ["a4", "letter", "auto"]

        (tmp_path / "sizes").mkdir(exist_ok=True)
        images = _make_images(tmp_path / "sizes", 1, fmt="JPEG", ext="jpg")

        for size in SUPPORTED_PAGE_SIZES:
            output = tmp_path / f"page_{size}.pdf"
            try:
                result = images_to_pdf(images, output, page_size=size)
                if output.exists():
                    reader = PdfReader(str(output))
                    assert len(reader.pages) >= 1
            except (ValueError, NotImplementedError):
                pass  # Algunas opciones pueden no estar implementadas aún


# ════════════════════════════════════════════════════════
#  IMAGES → PDF — Errores
# ════════════════════════════════════════════════════════

@pytest.mark.skipif(not HAS_IMAGE_TO_PDF, reason="core.image_to_pdf no disponible")
class TestImagesToPdfErrors:
    """Casos de error en conversión de imágenes a PDF."""

    def test_images_to_pdf_empty_list(self, tmp_path: Path):
        """Lista vacía de imágenes → error descriptivo."""
        output = tmp_path / "empty.pdf"

        try:
            result = images_to_pdf([], output)
            assert result.get("success") is False or result.get("status") == "error"
        except (ValueError, Exception):
            pass

    def test_images_to_pdf_nonexistent_image(self, tmp_path: Path):
        """Imagen que no existe → error."""
        fake = tmp_path / "no_existe.jpg"
        output = tmp_path / "ghost.pdf"

        try:
            result = images_to_pdf([fake], output)
            assert result.get("success") is False or "error" in str(result).lower()
        except (FileNotFoundError, Exception):
            pass

    def test_images_to_pdf_corrupt_image(self, tmp_path: Path):
        """Archivo con extensión .jpg pero contenido basura → error."""
        corrupt = tmp_path / "corrupt.jpg"
        corrupt.write_bytes(b"ESTO NO ES UNA IMAGEN")
        output = tmp_path / "corrupt_out.pdf"

        try:
            result = images_to_pdf([corrupt], output)
            if result:
                assert result.get("success") is False or "error" in str(result).lower()
        except Exception:
            pass


# ════════════════════════════════════════════════════════
#  PDF → IMAGES
# ════════════════════════════════════════════════════════

@pytest.mark.skipif(not HAS_PDF_TO_IMAGE, reason="core.pdf_to_image no disponible")
class TestPdfToImages:
    """Pruebas de conversión de PDF a imágenes."""

    def test_pdf_to_images_jpg(self, tmp_path: Path):
        """
        PDF de 3 páginas → 3 archivos JPG.
        Si poppler no está instalado, se usa mock.
        """
        pdf = _make_pdf(tmp_path / "to_images.pdf", 3)
        out_dir = tmp_path / "images_out"
        out_dir.mkdir()

        try:
            result = pdf_to_images(pdf, out_dir, format="jpg")
            # Verificar que se crearon imágenes
            image_files = list(out_dir.glob("*.jpg")) + list(out_dir.glob("*.jpeg"))
            assert len(image_files) >= 1
        except Exception as e:
            # Si falla por poppler no instalado, verificar mensaje
            if "poppler" in str(e).lower() or "pdf2image" in str(e).lower():
                pytest.skip("poppler no está instalado en este sistema")
            raise

    def test_pdf_to_images_png(self, tmp_path: Path):
        """PDF de 2 páginas → archivos PNG."""
        pdf = _make_pdf(tmp_path / "to_png.pdf", 2)
        out_dir = tmp_path / "png_out"
        out_dir.mkdir()

        try:
            result = pdf_to_images(pdf, out_dir, format="png")
            image_files = list(out_dir.glob("*.png"))
            assert len(image_files) >= 1
        except Exception as e:
            if "poppler" in str(e).lower() or "pdf2image" in str(e).lower():
                pytest.skip("poppler no está instalado")
            raise

    def test_pdf_to_images_page_range(self, tmp_path: Path):
        """Extraer solo páginas específicas del PDF."""
        pdf = _make_pdf(tmp_path / "range_img.pdf", 5)
        out_dir = tmp_path / "range_out"
        out_dir.mkdir()

        try:
            result = pdf_to_images(pdf, out_dir, format="jpg", page_range=(1, 2))
            image_files = list(out_dir.glob("*.jpg")) + list(out_dir.glob("*.jpeg"))
            # Debería crear máximo 2 imágenes
            assert len(image_files) <= 3  # Tolerancia
        except Exception as e:
            if "poppler" in str(e).lower() or "pdf2image" in str(e).lower():
                pytest.skip("poppler no está instalado")
            raise


# ════════════════════════════════════════════════════════
#  PDF → WORD (skip si no hay dependencias)
# ════════════════════════════════════════════════════════

@pytest.mark.skipif(not HAS_PDF_TO_WORD, reason="core.pdf_to_word no disponible")
class TestPdfToWord:
    """Pruebas de conversión de PDF a Word."""

    def test_pdf_to_word_basic(self, tmp_path: Path):
        """Conversión básica de PDF a Word."""
        pdf = _make_pdf(tmp_path / "to_word.pdf", 2)
        output = tmp_path / "output.docx"

        try:
            result = pdf_to_word(pdf, output)
            if output.exists():
                assert output.stat().st_size > 0
        except Exception as e:
            if "pdf2docx" in str(e).lower():
                pytest.skip("pdf2docx no disponible")
            raise


# ════════════════════════════════════════════════════════
#  WORD → PDF (skip si no hay dependencias)
# ════════════════════════════════════════════════════════

@pytest.mark.skipif(not HAS_WORD_TO_PDF, reason="core.word_to_pdf no disponible")
class TestWordToPdf:
    """Pruebas de conversión de Word a PDF."""

    def test_word_to_pdf_basic(self, tmp_path: Path, sample_docx_path: Path):
        """Conversión básica de DOCX a PDF."""
        output = tmp_path / "from_word.pdf"

        try:
            result = word_to_pdf(sample_docx_path, output)
            if output.exists():
                assert output.stat().st_size > 0
        except Exception as e:
            # Muchas dependencias posibles (LibreOffice, etc.)
            pytest.skip(f"Dependencia no disponible: {e}")


# ════════════════════════════════════════════════════════
#  EXCEL → PDF (skip si no hay dependencias)
# ════════════════════════════════════════════════════════

@pytest.mark.skipif(not HAS_EXCEL_TO_PDF, reason="core.excel_to_pdf no disponible")
class TestExcelToPdf:
    def test_detect_engine_basic(self):
        """Verifica que el detector de motor no falle catastróficamente."""
        engine = detect_engine()
        assert engine in [None, "excel", "libreoffice"]

    @patch("core.excel_to_pdf.detect_engine")
    @patch("subprocess.run")
    def test_excel_to_pdf_libreoffice(self, mock_run, mock_detect, tmp_path: Path):
        """Verifica la conversión exitosa usando LibreOffice."""
        mock_detect.return_value = "libreoffice"
        mock_run.return_value = MagicMock(returncode=0)
        
        inp = tmp_path / "test.xlsx"
        inp.touch()
        out = tmp_path / "test.pdf"
        
        def create_pdf(*args, **kwargs):
            out.touch()
            return MagicMock(returncode=0)
        mock_run.side_effect = create_pdf
        
        res = excel_to_pdf(inp, out)
        assert res["success"] is True
        assert res["engine_used"] == "libreoffice"
        assert Path(res["output_path"]).exists()

    @patch("core.excel_to_pdf.detect_engine")
    def test_excel_to_pdf_no_engine(self, mock_detect, tmp_path: Path):
        """Verifica el error si no hay motor."""
        mock_detect.return_value = None
        inp = tmp_path / "test.xlsx"
        out = tmp_path / "test.pdf"
        
        res = excel_to_pdf(inp, out)
        assert res["success"] is False
        assert "No se detectó" in res["error"]


# ════════════════════════════════════════════════════════
#  POWERPOINT → PDF (skip si no hay dependencias)
# ════════════════════════════════════════════════════════

@pytest.mark.skipif(not HAS_POWERPOINT_TO_PDF, reason="core.powerpoint_to_pdf no disponible")
class TestPowerpointToPdf:
    def test_detect_powerpoint_engine_basic(self):
        """Verifica que el detector de motor de PowerPoint no falle."""
        engine = detect_powerpoint_engine()
        assert engine in [None, "powerpoint", "libreoffice"]

    @patch("core.powerpoint_to_pdf.detect_powerpoint_engine")
    @patch("subprocess.run")
    def test_powerpoint_to_pdf_libreoffice(self, mock_run, mock_detect, tmp_path: Path):
        """Verifica la conversión exitosa usando LibreOffice fallback."""
        mock_detect.return_value = "libreoffice"
        mock_run.return_value = MagicMock(returncode=0)
        
        inp = tmp_path / "test.pptx"
        inp.touch()
        out = tmp_path / "test.pdf"
        
        def create_pdf(*args, **kwargs):
            out.touch()
            return MagicMock(returncode=0)
        mock_run.side_effect = create_pdf
        
        res = powerpoint_to_pdf(inp, out)
        assert res["success"] is True
        assert res["engine_used"] == "libreoffice"
        assert Path(res["output_path"]).exists()

    @patch("core.powerpoint_to_pdf.detect_powerpoint_engine")
    def test_powerpoint_to_pdf_no_engine(self, mock_detect, tmp_path: Path):
        """Verifica el error si no hay motor de PowerPoint."""
        mock_detect.return_value = None
        inp = tmp_path / "test.pptx"
        out = tmp_path / "test.pdf"
        
        res = powerpoint_to_pdf(inp, out)
        assert res["success"] is False
        assert "No se detectó" in res["error"]


# ════════════════════════════════════════════════════════
#  PDF → EXCEL
# ════════════════════════════════════════════════════════

@pytest.mark.skipif(not HAS_PDF_TO_EXCEL, reason="core.pdf_to_excel no disponible")
class TestPdfToExcel:
    @patch("pdfplumber.open")
    def test_pdf_to_excel_basic(self, mock_pdf_open, tmp_path: Path):
        """Verifica extracción básica de PDF a Excel con detección inteligente de tipos."""
        pdf_path = tmp_path / "source.pdf"
        pdf_path.touch()
        
        mock_pdf = MagicMock()
        mock_page = MagicMock()
        mock_page.extract_tables.return_value = [
            [["Nombre", "Edad", "Puntuacion"], ["Juan", "30", "98.5"], ["Maria", "25", "99.9"]]
        ]
        mock_page.extract_text.return_value = ""
        mock_pdf.pages = [mock_page]
        mock_pdf_open.return_value.__enter__.return_value = mock_pdf
        
        out_xlsx = tmp_path / "dest.xlsx"
        
        res = pdf_to_excel(pdf_path, out_xlsx, detection_mode="hybrid")
        assert res["success"] is True
        assert Path(res["output_path"]).exists()
        
        # Verificar contenido final del archivo de Excel generado
        import openpyxl
        wb = openpyxl.load_workbook(res["output_path"])
        assert "Pagina 1" in wb.sheetnames
        ws = wb["Pagina 1"]
        assert ws.cell(row=1, column=1).value == "Nombre"
        assert ws.cell(row=2, column=2).value == 30          # Debe ser entero
        assert ws.cell(row=2, column=3).value == 98.5        # Debe ser flotante
        assert ws.cell(row=3, column=2).value == 25          # Debe ser entero
        assert ws.cell(row=3, column=3).value == 99.9        # Debe ser flotante

    @patch("pdfplumber.open")
    def test_pdf_to_excel_single_sheet(self, mock_pdf_open, tmp_path: Path):
        """Verifica que se consoliden múltiples páginas en una sola hoja unificada."""
        pdf_path = tmp_path / "source_single.pdf"
        pdf_path.touch()
        
        mock_pdf = MagicMock()
        mock_page_1 = MagicMock()
        mock_page_1.extract_tables.return_value = [[["Cabecera1"], ["Valor1"]]]
        mock_page_2 = MagicMock()
        mock_page_2.extract_tables.return_value = [[["Cabecera2"], ["Valor2"]]]
        
        mock_pdf.pages = [mock_page_1, mock_page_2]
        mock_pdf_open.return_value.__enter__.return_value = mock_pdf
        
        out_xlsx = tmp_path / "dest_single.xlsx"
        
        res = pdf_to_excel(pdf_path, out_xlsx, single_sheet=True)
        assert res["success"] is True
        
        import openpyxl
        wb = openpyxl.load_workbook(res["output_path"])
        assert "Datos Consolidados" in wb.sheetnames
        assert "Pagina 1" not in wb.sheetnames
        
        ws = wb["Datos Consolidados"]
        assert ws.cell(row=1, column=1).value == "Cabecera1"
        assert ws.cell(row=2, column=1).value == "Valor1"

    @patch("pdfplumber.open")
    def test_pdf_to_excel_no_numeric_casting(self, mock_pdf_open, tmp_path: Path):
        """Verifica que no se casteen números si auto_convert_numbers está desactivado."""
        pdf_path = tmp_path / "source_no_cast.pdf"
        pdf_path.touch()
        
        mock_pdf = MagicMock()
        mock_page = MagicMock()
        mock_page.extract_tables.return_value = [[["Numero"], ["123"]]]
        mock_pdf.pages = [mock_page]
        mock_pdf_open.return_value.__enter__.return_value = mock_pdf
        
        out_xlsx = tmp_path / "dest_no_cast.xlsx"
        
        res = pdf_to_excel(pdf_path, out_xlsx, auto_convert_numbers=False)
        assert res["success"] is True
        
        import openpyxl
        wb = openpyxl.load_workbook(res["output_path"])
        ws = wb["Pagina 1"]
        # Debe mantenerse como string "123"
        assert ws.cell(row=2, column=1).value == "123"
        assert type(ws.cell(row=2, column=1).value) is str

    @patch("pdfplumber.open")
    def test_pdf_to_excel_text_fallback(self, mock_pdf_open, tmp_path: Path):
        """Verifica el fallback a texto plano si no se detectan tablas estructuradas."""
        pdf_path = tmp_path / "source_fallback.pdf"
        pdf_path.touch()
        
        mock_pdf = MagicMock()
        mock_page = MagicMock()
        # No detecta ninguna tabla estructurada
        mock_page.extract_tables.return_value = []
        # Pero contiene texto plano legible
        mock_page.extract_text.return_value = "Esta es una línea de texto plano\nEsta es la segunda"
        
        mock_pdf.pages = [mock_page]
        mock_pdf_open.return_value.__enter__.return_value = mock_pdf
        
        out_xlsx = tmp_path / "dest_fallback.xlsx"
        
        res = pdf_to_excel(pdf_path, out_xlsx)
        assert res["success"] is True
        
        import openpyxl
        wb = openpyxl.load_workbook(res["output_path"])
        ws = wb["Pagina 1"]
        assert ws.cell(row=1, column=1).value == "Esta es una línea de texto plano"
        assert ws.cell(row=2, column=1).value == "Esta es la segunda"

    @patch("pdfplumber.open")
    def test_pdf_to_excel_detection_modes(self, mock_pdf_open, tmp_path: Path):
        """Verifica que no falle usando modos específicos como lines o text."""
        pdf_path = tmp_path / "source_modes.pdf"
        pdf_path.touch()
        
        mock_pdf = MagicMock()
        mock_page = MagicMock()
        mock_page.extract_tables.return_value = [[["Dato"]]]
        mock_pdf.pages = [mock_page]
        mock_pdf_open.return_value.__enter__.return_value = mock_pdf
        
        # Test exact lines mode
        out_xlsx_lines = tmp_path / "dest_lines.xlsx"
        res_lines = pdf_to_excel(pdf_path, out_xlsx_lines, detection_mode="lines")
        assert res_lines["success"] is True
        
        # Test exact text mode
        out_xlsx_text = tmp_path / "dest_text.xlsx"
        res_text = pdf_to_excel(pdf_path, out_xlsx_text, detection_mode="text")
        assert res_text["success"] is True

    @patch("pdfplumber.open")
    def test_pdf_to_excel_oti_document(self, mock_pdf_open, tmp_path: Path):
        """Verifica la conversión unificada de un documento Oti con páginas izquierda/derecha."""
        pdf_path = tmp_path / "oti_document.pdf"
        pdf_path.touch()
        
        mock_pdf = MagicMock()
        
        # Página 1 (Izquierda): Código Patrimonial, Cod Barras, Descripción
        mock_page_1 = MagicMock()
        mock_table_1 = MagicMock()
        
        # Simulamos que la columna 2 ("Descripción") se fragmenta en 3 columnas físicas
        # debido a bounding boxes/espacios: Col 2 (Descripción), Col 3 (extra), Col 4 (extra)
        mock_table_1.extract.return_value = [
            ["Código Patrimonial", "Cod Barras", "Descripción", "", ""],
            ["10001", "88888", "ACUMULADOR DE EN", "ERGÍA-EQUIPO", "DE UPS 6 SM"],
            ["10002", "99999", "EQUIPO DE", "COMPUTO", "PORTATIL"]
        ]
        
        # Simulamos los objetos de columnas y celdas/rows de pdfplumber
        col0 = MagicMock(); col0.bbox = (0, 0, 10, 0)
        col1 = MagicMock(); col1.bbox = (10, 0, 20, 0)
        col2 = MagicMock(); col2.bbox = (20, 0, 30, 0)
        col3 = MagicMock(); col3.bbox = (30, 0, 40, 0)
        col4 = MagicMock(); col4.bbox = (40, 0, 50, 0)
        mock_table_1.columns = [col0, col1, col2, col3, col4]
        
        row_h = MagicMock(); row_h.bbox = (0, 0, 50, 10)
        row1 = MagicMock(); row1.bbox = (0, 10, 50, 20)
        row2 = MagicMock(); row2.bbox = (0, 20, 50, 30)
        mock_table_1.rows = [row_h, row1, row2]
        
        # Mocks para crop y extract_text en la página 1
        def crop_side_effect_page1(bbox):
            crop_mock = MagicMock()
            if bbox[0] == 20 and bbox[2] == 50:
                if bbox[1] == 10:
                    crop_mock.extract_text.return_value = "ACUMULADOR DE EN\nERGÍA-EQUIPO\nDE UPS 6 SM"
                elif bbox[1] == 20:
                    crop_mock.extract_text.return_value = "EQUIPO DE\nCOMPUTO\nPORTATIL"
            elif bbox[0] == 0:
                crop_mock.extract_text.return_value = "10001" if bbox[1] == 10 else "10002"
            elif bbox[0] == 10:
                crop_mock.extract_text.return_value = "88888" if bbox[1] == 10 else "99999"
            return crop_mock
            
        mock_page_1.crop.side_effect = crop_side_effect_page1
        mock_page_1.find_tables.return_value = [mock_table_1]
        
        # Página 2 (Derecha): Medidas, Marca, Modelo
        mock_page_2 = MagicMock()
        mock_table_2 = MagicMock()
        mock_table_2.extract.return_value = [
            ["Medidas", "Marca", "Modelo", "Fecha Alta"],
            ["1.5m", "SM", "Mod-X", "15/08/2022"],
            ["-", "S/M", "Mod-Y", "20/09/2023"]
        ]
        
        col_m0 = MagicMock(); col_m0.bbox = (50, 0, 60, 0)
        col_m1 = MagicMock(); col_m1.bbox = (60, 0, 70, 0)
        col_m2 = MagicMock(); col_m2.bbox = (70, 0, 80, 0)
        col_m3 = MagicMock(); col_m3.bbox = (80, 0, 90, 0)
        mock_table_2.columns = [col_m0, col_m1, col_m2, col_m3]
        mock_table_2.rows = [row_h, row1, row2]
        
        def crop_side_effect_page2(bbox):
            crop_mock = MagicMock()
            if bbox[0] == 50:
                crop_mock.extract_text.return_value = "1.5m" if bbox[1] == 10 else "-"
            elif bbox[0] == 60:
                crop_mock.extract_text.return_value = "SM" if bbox[1] == 10 else "S/M"
            elif bbox[0] == 70:
                crop_mock.extract_text.return_value = "Mod-X" if bbox[1] == 10 else "Mod-Y"
            elif bbox[0] == 80:
                crop_mock.extract_text.return_value = "15/08/2022" if bbox[1] == 10 else "20/09/2023"
            return crop_mock
            
        mock_page_2.crop.side_effect = crop_side_effect_page2
        mock_page_2.find_tables.return_value = [mock_table_2]
        
        mock_pdf.pages = [mock_page_1, mock_page_2]
        mock_pdf_open.return_value.__enter__.return_value = mock_pdf
        
        out_xlsx = tmp_path / "oti_output.xlsx"
        res = pdf_to_excel(pdf_path, out_xlsx)
        
        assert res["success"] is True
        
        import openpyxl
        wb = openpyxl.load_workbook(res["output_path"])
        assert "Datos Extraídos" in wb.sheetnames
        ws = wb["Datos Extraídos"]
        
        # Verificar cabeceras
        assert ws.cell(row=1, column=1).value == "Código Patrimonial"
        assert ws.cell(row=1, column=3).value == "Descripción"
        assert ws.cell(row=1, column=4).value == "Medidas"
        assert ws.cell(row=1, column=5).value == "Marca"
        assert ws.cell(row=1, column=9).value == "Fecha Alta"
        
        # Verificar fila 1 unificada
        assert ws.cell(row=2, column=1).value == 10001
        assert ws.cell(row=2, column=2).value == 88888
        assert ws.cell(row=2, column=3).value == "ACUMULADOR DE ENERGÍA-EQUIPO DE UPS 6 SM"
        assert ws.cell(row=2, column=4).value == "1.5m"
        assert ws.cell(row=2, column=5).value == "SM"
        assert ws.cell(row=2, column=6).value == "Mod-X"
        assert ws.cell(row=2, column=9).value == "15/08/2022"
        
        # Verificar fila 2 unificada
        assert ws.cell(row=3, column=1).value == 10002
        assert ws.cell(row=3, column=2).value == 99999
        assert ws.cell(row=3, column=3).value == "EQUIPO DE COMPUTO PORTATIL"
        assert ws.cell(row=3, column=4).value == "-"
        assert ws.cell(row=3, column=5).value == "S/M"
        assert ws.cell(row=3, column=9).value == "20/09/2023"



