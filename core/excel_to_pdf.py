import logging
import os
import subprocess
from pathlib import Path
from typing import Callable, Optional
import winreg

logger = logging.getLogger(__name__)

def detect_engine() -> str | None:
    """Detecta si Microsoft Excel, LibreOffice o ReportLab están disponibles."""
    from utils.validators import check_office_installed
    
    # 1. Chequear Excel via COM Dispatch
    if os.name == 'nt' and check_office_installed("Excel.Application"):
        return "excel"
        
    # 2. Chequear LibreOffice
    libreoffice_paths = []
    if os.name == 'nt':
        libreoffice_paths = [
            Path(os.environ.get("PROGRAMFILES", "C:\\Program Files")) / "LibreOffice" / "program" / "soffice.exe",
            Path(os.environ.get("PROGRAMFILES(X86)", "C:\\Program Files (x86)")) / "LibreOffice" / "program" / "soffice.exe",
        ]
    
    for path in libreoffice_paths:
        if path.exists():
            return "libreoffice"
            
    # Chequear PATH para LibreOffice
    try:
        if subprocess.run(["soffice", "--version"], capture_output=True).returncode == 0:
            return "libreoffice"
    except Exception:
        pass
        
    # 3. Chequear ReportLab
    try:
        import reportlab
        import openpyxl
        return "reportlab"
    except ImportError:
        pass
        
    return None

def _safe_output(output_path: Path) -> Path:
    """Genera una ruta segura añadiendo sufijos si el archivo ya existe."""
    final_output = output_path
    counter = 1
    while final_output.exists():
        final_output = output_path.parent / f"{output_path.stem}_{counter}{output_path.suffix}"
        counter += 1
    return final_output

def _convert_excel_com(input_path: Path, output_path: Path, progress_callback: Optional[Callable[[int, str], None]]) -> None:
    """Convierte usando Microsoft Excel vía automatización COM (solo Windows)."""
    import win32com.client
    import pythoncom
    
    if progress_callback:
        progress_callback(30, "Abriendo Microsoft Excel en segundo plano...")
        
    pythoncom.CoInitialize()
    excel = None
    wb = None
    try:
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        
        if progress_callback:
            progress_callback(50, "Cargando libro de Excel...")
            
        wb = excel.Workbooks.Open(str(input_path.absolute()))
        
        # Ajustar diseño de página para mejor presentación (Apaisado, centrado, 1 página de ancho)
        for ws in wb.Worksheets:
            try:
                ws.PageSetup.Orientation = 2 # 2 = xlLandscape
                ws.PageSetup.Zoom = False
                ws.PageSetup.FitToPagesWide = 1
                ws.PageSetup.FitToPagesTall = False
                ws.PageSetup.CenterHorizontally = True
            except Exception as e:
                logger.warning(f"No se pudo ajustar diseño de {ws.Name}: {e}")
        
        if progress_callback:
            progress_callback(75, "Exportando hojas a formato PDF...")
            
        # 0 es xlTypePDF
        wb.ExportAsFixedFormat(0, str(output_path.absolute()))
        
        if progress_callback:
            progress_callback(95, "Cerrando Microsoft Excel...")
            
        wb.Close(False)
    finally:
        if excel:
            excel.Quit()
        pythoncom.CoUninitialize()

def _convert_libreoffice(input_path: Path, output_path: Path, progress_callback: Optional[Callable[[int, str], None]]) -> None:
    """Convierte usando LibreOffice headless."""
    if progress_callback:
        progress_callback(30, "Ejecutando LibreOffice en segundo plano...")
        
    cmd = [
        "soffice",
        "--headless",
        "--convert-to", "pdf",
        "--outdir", str(output_path.parent.absolute()),
        str(input_path.absolute())
    ]
    
    if os.name == 'nt' and not subprocess.run(["where", "soffice"], capture_output=True).returncode == 0:
        libreoffice_path = Path(os.environ.get("PROGRAMFILES", "C:\\Program Files")) / "LibreOffice" / "program" / "soffice.exe"
        if libreoffice_path.exists():
            cmd[0] = str(libreoffice_path)
            
    process = subprocess.run(cmd, capture_output=True, text=True)
    if process.returncode != 0:
        raise Exception(f"Error de LibreOffice: {process.stderr}")
        
    lo_output = output_path.parent / f"{input_path.stem}.pdf"
    if lo_output.exists() and lo_output != output_path:
        if output_path.exists():
            output_path.unlink()
        lo_output.rename(output_path)

def _convert_reportlab(input_path: Path, output_path: Path, progress_callback: Optional[Callable[[int, str], None]]) -> None:
    """Convierte usando Python puro (ReportLab + openpyxl)."""
    if progress_callback:
        progress_callback(20, "Cargando librerías de generación PDF (ReportLab)...")
        
    import openpyxl
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak
    from reportlab.lib import colors
    
    if progress_callback:
        progress_callback(40, "Leyendo archivo Excel...")
        
    wb = openpyxl.load_workbook(input_path, data_only=True)
    
    doc = SimpleDocTemplate(str(output_path), pagesize=landscape(A4), 
                            rightMargin=20, leftMargin=20, 
                            topMargin=20, bottomMargin=20)
    elements = []
    
    total_sheets = len(wb.sheetnames)
    for idx, sheet_name in enumerate(wb.sheetnames):
        if progress_callback:
            progress_callback(40 + int((idx/total_sheets)*40), f"Procesando hoja: {sheet_name}...")
            
        ws = wb[sheet_name]
        data = []
        max_col = 0
        for row in ws.iter_rows(values_only=True):
            # Limpiar celdas None
            clean_row = [str(cell) if cell is not None else "" for cell in row]
            if any(clean_row): # Si no está toda la fila vacía
                data.append(clean_row)
                max_col = max(max_col, len(clean_row))
                
        if not data:
            continue
            
        # Homogeneizar longitudes de fila
        for row in data:
            while len(row) < max_col:
                row.append("")
                
        # Anchos proporcionales básicos
        col_widths = [0] * max_col
        for row in data:
            for c_idx, cell_str in enumerate(row):
                col_widths[c_idx] = max(col_widths[c_idx], len(str(cell_str)))
        
        # Ajustar para que encaje en la página A4 landscape (~800 puntos utiles)
        total_chars = sum(col_widths) or 1
        page_width = landscape(A4)[0] - 40
        widths = [max((w / total_chars) * page_width, 20) for w in col_widths]
        
        t = Table(data, colWidths=widths)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.grey),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('BOTTOMPADDING', (0,0), (-1,0), 12),
            ('BACKGROUND', (0,1), (-1,-1), colors.white),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('FONTNAME', (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE', (0,1), (-1,-1), 8),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('WORDWRAP', (0,0), (-1,-1), 'CJK')
        ]))
        
        elements.append(t)
        if idx < total_sheets - 1:
            elements.append(PageBreak())
            
    if progress_callback:
        progress_callback(85, "Generando archivo PDF final...")
        
    doc.build(elements)

def excel_to_pdf(
    input_path: Path,
    output_path: Path,
    preferred_engine: str = None,
    progress_callback: Optional[Callable[[int, str], None]] = None
) -> dict:
    """
    Convierte Excel a PDF usando el motor especificado o el mejor disponible.
    """
    result = {
        "success": False,
        "output_path": None,
        "engine_used": None,
        "error": None
    }
    
    # Detección de motor
    engine = preferred_engine
    if not engine:
        engine = detect_engine()
        
    if not engine:
        result["error"] = "No se detectó ningún motor disponible (Excel, LibreOffice o ReportLab)."
        return result
        
    result["engine_used"] = engine
    
    final_output = _safe_output(output_path)
        
    try:
        if engine == "excel":
            _convert_excel_com(input_path, final_output, progress_callback)
        elif engine == "libreoffice":
            _convert_libreoffice(input_path, final_output, progress_callback)
        elif engine == "reportlab":
            _convert_reportlab(input_path, final_output, progress_callback)
        else:
            raise ValueError(f"Motor '{engine}' no soportado.")

        if progress_callback:
            progress_callback(100, "¡Conversión finalizada!")

        result["success"] = True
        result["output_path"] = str(final_output)

    except PermissionError:
        result["error"] = "No se pudo escribir el archivo. Verifica que no esté abierto en otro programa."
    except Exception as e:
        logger.error(f"Error convirtiendo Excel a PDF ({engine}): {e}")
        result["error"] = f"Error en la conversión: {e}"
        
    return result
